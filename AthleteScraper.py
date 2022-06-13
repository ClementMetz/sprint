import time
from datetime import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from auxiliary import athle_regressor,clean_up_perf,standardize_event
import numpy as np
import openpyxl
import argparse


def parse_option():
    parser = argparse.ArgumentParser(description='Scrape athlete data')
    parser.add_argument('--workdir', type=str,default='ISPInput.xlsx',help='Workbook directory')
    parser.add_argument('--fill_perf',action='store_true',help='Specify in order to automatically fill PerfM and PerfW')
    args = parser.parse_args()
    return(args)


def initialize(athletes_table):
    #print(athletes_table)
    ath_table = []
    for i in range(2,athletes_table.max_row+1):
        if athletes_table.cell(row=i,column=1).value==None:
            break
        ath = {}
        licence = athletes_table.cell(row = i,column = 3).value
        #print(licence)
        if licence == None:
            ath['licence'] = ''
        else:
            ath['licence'] = str(int(athletes_table.cell(row = i,column = 3).value))
        ath['name'] = athletes_table.cell(row=i,column=1).value
        ath['firstname'] = athletes_table.cell(row=i,column=2).value
        ath['gender'] = athletes_table.cell(row=i,column=4).value
        print(ath) #ath = [nom,prenom,licence,gender]
        ath_table.append(ath)
    return(ath_table)

def request(driver,athletename,firstname,gender,by_licence_nb=False,licence_nb=0,clubname=''):
    print('*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*')
    now = datetime.now().year
    regressor = athle_regressor()
    element = driver.find_element(By.XPATH,'/html/body/div/div[2]/table/tbody/tr/td[4]')
    element.click()
    last_SBs = {}

    for y in range(now-10,now+1):
        current_SBs = {}
        year = str(y)
        
        select = Select(driver.find_element(By.XPATH,'/html/body/div/div[2]/div[5]/div/form/table/tbody/tr/td/table/tbody/tr[1]/td[2]/select'))
        select.select_by_visible_text(year)
        
        if by_licence_nb: #to avoid namesakes
            element = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[5]/div/form/table/tbody/tr/td/table/tbody/tr[6]/td[2]/input')
            element.send_keys(licence_nb)

        element = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[5]/div/form/table/tbody/tr/td/table/tbody/tr[3]/td[2]/input')
        element.send_keys(athletename)
        #time.sleep(0.2)
        element = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[5]/div/form/table/tbody/tr/td/table/tbody/tr[4]/td[2]/input')
        element.send_keys(firstname)
        #time.sleep(0.2)
        element = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[5]/div/form/div[1]/input')
        element.click()
        #time.sleep(0.5)
        tablexpath = '/html/body/div/div[2]/table[2]/tbody'

        try:
            table = driver.find_element(By.XPATH,tablexpath)
            soup = BeautifulSoup(table.get_attribute('innerHTML'), 'html.parser')
            t = soup.get_text(separator='|||').split('\n')
            for i in range (2,len(t)):
                l = t[i].split('|||')
                #print(l)
                date = l[1]
                event = l[3]
                event,hidden_event = standardize_event(event,gender)
                lh = len(hidden_event)
                if event[lh-1] == 'i': #indoor events are outdoor events
                    event = event[:lh-1]
                perf = l[6]
                perf = clean_up_perf(perf,event)

                if event in ['100mM','200mM','400mM','800mM','1500mM','3000mM','3000mscM','110mHM','400mHM','5000walkM','longM','highM','tripleM','poleM',
                'shotM','javM','hammerM','discM','100mW','200mW','400mW','800mW','1500mW','3000mW','100mHW','400mHW','3000walkW','longW','highW','tripleW','poleW','shotW','javW','hammerW','discW']:
                    new = regressor.reg(hidden_event,perf)
                    if new>0 and (event not in current_SBs.keys() or new > regressor.reg(hidden_event,current_SBs[event])):
                        current_SBs[event] = perf
                

        except:
            pass

        for key in current_SBs.keys():
            #print(key)
            last_SBs[key] = current_SBs[key]
        
        driver.back()

    return(last_SBs)


def make_csv(dico,gender,wb,row,countM,countW,sM,sW,opt):
    sheet = wb['AthleteScraper']

    if opt.fill_perf:
        perf_sheet = wb["Perf"+gender]
        c = eval("count"+gender)
        s = eval("s"+gender)
  
    if gender=='W':
        list_events = ['100mW','200mW','400mW','800mW','1500mW','3000mW','100mHW','400mHW','3000walkW','longW','highW','tripleW','poleW','shotW',
                'javW','hammerW','discW']
    else:
        list_events = ['100mM','200mM','400mM','800mM','1500mM','3000mM','3000scM','110mHM','400mHM','5000walkM','longM','highM','tripleM','poleM',
                'shotM','javM','hammerM','discM']
    for i in range(len(list_events)):
        x = list_events[i]
        sheet.cell(column = i+6,row = row, value='')
        if opt.fill_perf:
            alias = sheet.cell(column = 5,row = row).value
            perf_sheet.cell(column=1, row = s+c, value=alias)
        if x in dico.keys():
            sheet.cell(column = i+6,row = row, value=dico[x])
            if opt.fill_perf:
                perf_sheet.cell(column=i+2, row = s+c, value=dico[x])

    


def main():
    opt = parse_option()
    wb = openpyxl.load_workbook(filename = opt.workdir)
    sheet = wb['AthleteScraper']
    ath_table = initialize(sheet)

    driver_options = webdriver.ChromeOptions()
    driver_options.add_argument("headless")
    driver_options.add_argument("disable-gpu")
    driver_options.add_argument("disable-extensions")
    driver_options.add_argument("no-sandbox")
    driver_options.add_argument("disable-dev-shm-usage")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=driver_options)
    url = "https://bases.athle.fr/asp.net/accueil.aspx?frmbase=resultats"
    driver.get(url)
    countM,countW = 0,0

    sM,sW=1,1
    if opt.fill_perf: #initialize sM,sW
        perfM_sheet = wb['PerfM']
        perfW_sheet = wb['PerfW']
        
        while True:
            if perfM_sheet.cell(row=sM,column=1).value==None: #locate first empty row
                break
            sM+=1
        while True:
            if perfW_sheet.cell(row=sW,column=1).value==None: #locate first empty row
                break
            sW+=1

    for row,ath in enumerate(ath_table):
        last_SBs = request(driver,ath['name'],ath['firstname'],ath['gender'],True,ath['licence'],'')
        print(ath['firstname'],ath['name'])
        print(last_SBs)
        make_csv(last_SBs,ath['gender'],wb,row+2,countM,countW,sM,sW,opt)

        if ath['gender']=='M':
            countM+=1
        else:
            countW+=1
        
    wb.save(filename = opt.workdir)


if __name__ == '__main__':
    main()