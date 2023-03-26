import time
from datetime import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from auxiliary import athle_regressor,clean_up_perf,standardize_event
import numpy as np
import openpyxl
import argparse
import os
import warnings
from DrawAthleteInfo import request, date_to_float
from matplotlib import pyplot as plt
from numpy import polyfit,polyval,arange

def parse_option():
    parser = argparse.ArgumentParser(description='Scrape athlete data')
    parser.add_argument('--workdir', type=str,default='ISPInput.xlsx',help='Workbook directory')
    parser.add_argument('--fill_perf',action='store_true',help='Specify in order to automatically fill PerfM and PerfW')
    args = parser.parse_args()
    return(args)


def initialize(athletes_table):
    #print(athletes_table)
    ath_table = []
    for i in range(2,athletes_table.max_row+1):
        if athletes_table.cell(row=i,column=1).value==None:
            break
        ath = {}
        licence = athletes_table.cell(row = i,column = 3).value
        #print(licence)
        if licence == None:
            ath['licence'] = ''
        else:
            ath['licence'] = str(int(athletes_table.cell(row = i,column = 3).value))
        ath['name'] = athletes_table.cell(row=i,column=1).value
        ath['firstname'] = athletes_table.cell(row=i,column=2).value
        ath['gender'] = athletes_table.cell(row=i,column=4).value
        print(ath) #ath = [nom,prenom,licence,gender]
        ath_table.append(ath)
    return(ath_table)

def scrape_ath(driver,athletename,firstname,gender,by_licence_nb=False,licence_nb=0,clubname=''):
    print('*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*.*')
    data = request(driver,athletename,firstname,gender,by_licence_nb=False,licence_nb=0,clubname='')
    reg = athle_regressor()
    #print(data)
    year = datetime.now().year
    if gender=='W':
        list_events = ['100mW','200mW','400mW','800mW','1500mW','3000mW','100mHW','400mHW','3000walkW','longW','highW','tripleW','poleW','shotW',
                'javW','hammerW','discW','200mWi','400mWi','800mWi','1500mWi','3000mWi','longWi','highWi','tripleWi','poleWi','shotWi',
                'hammerWi','discWi']
    else:
        list_events = ['100mM','200mM','400mM','800mM','1500mM','3000mM','3000mscM','110mHM','400mHM','5000walkM','longM','highM','tripleM','poleM',
                'shotM','javM','hammerM','discM','200mMi','400mMi','800mMi','1500mMi','3000mMi','longMi','highMi','tripleMi','poleMi',
                'shotMi','hammerMi','discMi']
    clean_data = []
    for x in data:
        ev = x[1][0]
        if ev in list_events:
            clean_data.append(x)

    data_by_event = {}
    projected_perfs = {}

    for entry in clean_data:
        
        if type(entry[3])==int and entry[3]>0: 
            #print(entry)
            if entry[1][0][-1]=='i':
                indoor=True
                evinout = entry[1][0][:-1]
                bonus=' (i)'
            else:
                indoor=False
                evinout = entry[1][0]
                bonus=''
            if evinout not in data_by_event.keys():
                data_by_event[evinout] = [[entry[2]+bonus,int(entry[3]),int(entry[5])+date_to_float(entry[6])]]
            else:
                data_by_event[evinout].append([entry[2]+bonus,int(entry[3]),int(entry[5])+date_to_float(entry[6])])
            


    for event in data_by_event.keys():
        year_list = []
        perf_list = []
        label_list = []
        for x in data_by_event[event]:
            perf_list.append(x[1])
            year_list.append(x[2])
            label_list.append(x[0])
        
        last_year = max(year_list)
        
        fig, ax = plt.subplots()
        ax.set_title(event)
        ax.scatter(year_list, perf_list)
        if len(perf_list)>3: #avoid dumb plots
            if len(perf_list)>10:
                order=2
            else:
                order=1
            p = polyfit(year_list,perf_list,order)
            ax.plot(arange(min(year_list)-0.6,max(year_list)+0.6,0.02),[polyval(p,x) for x in arange(min(year_list)-0.6,max(year_list)+0.6,0.02)],'r')
            projected_points = polyval(p,last_year)
            projected_perfs[event] = reg.inv(event,projected_points)
            ax.annotate(projected_perfs[event], (last_year,projected_points))

        else:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                p = polyfit(year_list,perf_list,1)
            projected_points = polyval(p,last_year)
            projected_perfs[event] = reg.inv(event,projected_points)
        
        for i, txt in enumerate(label_list):
            ax.annotate(txt, (year_list[i], perf_list[i]))

        ax.scatter(last_year,projected_points,c='g')
        
        
        figdir = os.getcwd()+"/figures/"+athletename+'_'+firstname+'/'
        if not os.path.isdir(figdir):
            os.mkdir(figdir)
    
        plt.savefig(figdir+event+".png")

        
    #print(projected_perfs)

    return(projected_perfs)


def make_csv(dico,gender,wb,row,countM,countW,sM,sW,opt):
    sheet = wb['AthleteScraper']

    if opt.fill_perf:
        perf_sheet = wb["Perf"+gender]
        c = eval("count"+gender)
        s = eval("s"+gender)
  
    if gender=='W':
        list_events = ['100mW','200mW','400mW','800mW','1500mW','3000mW','100mHW','400mHW','3000walkW','longW','highW','tripleW','poleW','shotW',
                'javW','hammerW','discW']
    else:
        list_events = ['100mM','200mM','400mM','800mM','1500mM','3000mM','3000mscM','110mHM','400mHM','5000walkM','longM','highM','tripleM','poleM',
                'shotM','javM','hammerM','discM']
    for i in range(len(list_events)):
        x = list_events[i]
        sheet.cell(column = i+6,row = row, value='')
        if opt.fill_perf:
            alias = sheet.cell(column = 5,row = row).value
            perf_sheet.cell(column=1, row = s+c, value=alias)
        if x in dico.keys():
            sheet.cell(column = i+6,row = row, value=dico[x])
            if opt.fill_perf:
                perf_sheet.cell(column=i+2, row = s+c, value=dico[x])

    


def main():
    opt = parse_option()
    wb = openpyxl.load_workbook(filename = opt.workdir)
    sheet = wb['AthleteScraper']
    ath_table = initialize(sheet)

    driver_options = webdriver.ChromeOptions()
    driver_options.add_argument("headless")
    driver_options.add_argument("disable-gpu")
    driver_options.add_argument("disable-extensions")
    driver_options.add_argument("no-sandbox")
    driver_options.add_argument("disable-dev-shm-usage")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=driver_options)
    url = "https://bases.athle.fr/asp.net/accueil.aspx?frmbase=resultats"
    driver.get(url)
    countM,countW = 0,0

    sM,sW=1,1
    if opt.fill_perf: #initialize sM,sW
        perfM_sheet = wb['PerfM']
        perfW_sheet = wb['PerfW']
        
        while True:
            if perfM_sheet.cell(row=sM,column=1).value==None: #locate first empty row
                break
            sM+=1
        while True:
            if perfW_sheet.cell(row=sW,column=1).value==None: #locate first empty row
                break
            sW+=1

    for row,ath in enumerate(ath_table):
        last_SBs = scrape_ath(driver,ath['name'],ath['firstname'],ath['gender'],True,ath['licence'],'')
        print(ath['firstname'],ath['name'])
        print(last_SBs)
        make_csv(last_SBs,ath['gender'],wb,row+2,countM,countW,sM,sW,opt)

        if ath['gender']=='M':
            countM+=1
        else:
            countW+=1
        
    wb.save(filename = opt.workdir)


if __name__ == '__main__':
    main()